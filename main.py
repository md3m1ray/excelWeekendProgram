import tkinter
import tkinter as tk
from tkinter import ttk, Menu
import openpyxl

PATH = "people.xlsx"


def load_data():
    workbook = openpyxl.load_workbook(PATH)
    sheet = workbook.active
    list_values = list(sheet.values)

    for col_name in list_values[0]:
        treeview.heading(col_name, text=col_name)

    for value_tuple in list_values[1:]:
        treeview.insert('', tk.END, values=value_tuple)


def delete_row():
    workbook = openpyxl.load_workbook(PATH)
    sheet = workbook.active
    selected_item2 = treeview.focus()

    for row in sheet.iter_rows(min_row=2, min_col=1, max_row=100, max_col=1, values_only=False):
        for cell in row:
            if cell.value == f"{selected_item2}":
                sheet.delete_rows(cell.row)
                treeview.delete(selected_item2)

    workbook.save(PATH)
    get_all()


def insert_row():
    id_ = ""
    name = name_entry.get()
    day = status_combobox.get()

    workbook = openpyxl.load_workbook(PATH)
    sheet = workbook.active
    row_values = [id_, name, day]
    sheet.append(row_values)
    workbook.save(PATH)

    treeview.insert('', tk.END, values=row_values)

    name_entry.delete(0, "end")
    name_entry.insert(0, "İsim Soyisim")

    get_all()


def get_all():
    workbook = openpyxl.load_workbook(PATH)
    sheet = workbook.active
    row_num = 2

    for id_item in treeview.get_children():
        treeview.set(id_item, column=0, value=id_item)
        sheet.cell(row=row_num, column=1).value = id_item
        row_num += 1

    workbook.save(PATH)


def toggle_mode():
    if mode_switch.instate(["selected"]):
        style.theme_use("forest-light")
    else:
        style.theme_use("forest-dark")


def add_day():
    focused = treeview.focus()
    day = status_combobox.get()

    workbook = openpyxl.load_workbook(PATH)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, min_col=1, max_row=100, max_col=1, values_only=False):
        for cell in row:
            if cell.value == f"{focused}":
                sheet.cell(row=cell.row, column=3).value = day

    treeview.set(focused, column=2, value=day)

    workbook.save(PATH)


def saturday():
    focused = treeview.focus()
    day = "Cumartesi"

    workbook = openpyxl.load_workbook(PATH)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, min_col=1, max_row=100, max_col=1, values_only=False):
        for cell in row:
            if cell.value == f"{focused}":
                sheet.cell(row=cell.row, column=3).value = day

    treeview.set(focused, column=2, value=day)

    workbook.save(PATH)


def sunday():
    focused = treeview.focus()
    day = "Pazar"

    workbook = openpyxl.load_workbook(PATH)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, min_col=1, max_row=100, max_col=1, values_only=False):
        for cell in row:
            if cell.value == f"{focused}":
                sheet.cell(row=cell.row, column=3).value = day

    treeview.set(focused, column=2, value=day)

    workbook.save(PATH)


def day_popup(e):
    day_menu.tk_popup(e.x_root, e.y_root)


root = tk.Tk()

style = ttk.Style(root)
root.tk.call("source", "forest-light.tcl")
root.tk.call("source", "forest-dark.tcl")
style.theme_use("forest-dark")
root.title("Hafta Sonu Çalışma Çizelgesi")
combo_list = ["Cumartesi", "Pazar"]

frame = ttk.Frame(root)
frame.pack()

widgets_frame = ttk.LabelFrame(frame, text="Personel Ekle/Sil")
widgets_frame.grid(row=0, column=0, padx=20, pady=10)

name_entry = ttk.Entry(widgets_frame)
name_entry.insert(0, "İsim Soyisim")
name_entry.bind("<FocusIn>", lambda e: name_entry.delete('0', 'end'))
name_entry.grid(row=2, column=0, padx=5, pady=(0, 5), sticky="ew")

button = ttk.Button(widgets_frame, text="Personel Ekle", command=insert_row)
button.grid(row=3, column=0, padx=5, pady=5, sticky="nsew")

separator = ttk.Separator(widgets_frame)
separator.grid(row=4, column=0, padx=(20, 10), pady=10, sticky="ew")

button = ttk.Button(widgets_frame, text="Seç ve Sil", command=delete_row)
button.grid(row=6, column=0, padx=5, pady=5, sticky="nsew")

separator = ttk.Separator(widgets_frame)
separator.grid(row=7, column=0, padx=(20, 10), pady=10, sticky="ew")

separator = ttk.Separator(widgets_frame)
separator.grid(row=8, column=0, padx=(20, 10), pady=10, sticky="ew")

status_combobox = ttk.Combobox(widgets_frame, values=combo_list)
status_combobox.current(0)
status_combobox.grid(row=9, column=0, sticky="ew")

button = ttk.Button(widgets_frame, text="Seç ve Değiştir", command=add_day)
button.grid(row=10, column=0, padx=5, pady=5, sticky="nsew")

separator = ttk.Separator(widgets_frame)
separator.grid(row=11, column=0, padx=(20, 10), pady=10, sticky="ew")

mode_switch = ttk.Checkbutton(
    widgets_frame, text="Tema", style="Switch", command=toggle_mode)
mode_switch.grid(row=12, column=0, padx=5, pady=10, sticky="nsew")

treeFrame = ttk.Frame(frame)
treeFrame.grid(row=0, column=1, pady=10)
treeScroll = ttk.Scrollbar(treeFrame)
treeScroll.pack(side="right", fill="y")

cols = ("ID", "İsim Soyisim", "Gün")
treeview = ttk.Treeview(treeFrame, show="headings",
                        yscrollcommand=treeScroll.set, columns=cols, height=35)

treeview.column("ID", width=50, anchor=tkinter.CENTER)
treeview.column("İsim Soyisim", width=200, anchor=tkinter.CENTER)
treeview.column("Gün", width=65, anchor=tkinter.CENTER)

day_menu = Menu(root, tearoff=False)
day_menu.add_command(label="Cumartesi", command=saturday)
day_menu.add_command(label="Pazar", command=sunday)

root.bind("<Button-3>", day_popup)

treeview.pack()
treeScroll.config(command=treeview.yview)

load_data()
get_all()
root.mainloop()
